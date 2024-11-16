from tkinter import *
import tkinter as tk
from openpyxl import Workbook, load_workbook
import openpyxl
import pandas as pd
from datetime import date


def ltab():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    i = 0
    while True:
        i += 1
        if (sheet["A" + str(i)].value) == None:
            break
    return i - 1


def ltabm():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active
    i = 0
    while True:
        i += 1
        if (sheet["A" + str(i)].value) == None:
            break
    return i - 1


def ajout_pers():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    personne = {
        "cin": "",
        "nom": "",
        "prenom": "",
        "age": 1,
        "adresse": "",
        "nationalite": "",
        "tel": "",
        "date": [1, 1, 1],
        "décédé": 1,
    }
    for cle in personne:
        if cle != "date":
            print("donner ", cle, ":")
            personne[cle] = input()
        else:
            for i in range(3):
                if i == 0:
                    print("donner jour :")
                    personne["date"][i] = int(input())
                if i == 1:
                    print("donner mois :")
                    personne["date"][i] = int(input())
                if i == 2:
                    print("donner annee :")
                    personne["date"][i] = int(input())
    sheet.append(
        [
            personne["cin"],
            personne["nom"],
            personne["prenom"],
            personne["tel"],
            personne["nationalite"],
            personne["adresse"],
            personne["age"],
            personne["date"][0],
            personne["date"][1],
            personne["date"][2],
            personne["décédé"],
        ]
    )
    workbook.save("personnes.xlsx")
    exit()


def sup_pers():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    df = pd.read_excel("personnes.xlsx")

    def delt():
        cin = int(input("donner cin de personne désiré a supprimé :"))
        cin = str(cin)
        for i in range(1, ltab() + 1):
            if sheet["A" + str(i)].value == cin:
                sheet.delete_rows(i)

    delt()
    workbook.save("personnes.xlsx")
    exit()


def sup_nation():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active

    def delt():
        nation = str(input("donner nationalité de personnes désiré a supprimé :"))
        for i in range(1, ltab() + 1):
            if sheet["E" + str(i)].value == nation:
                sheet.delete_rows(i)

    delt()
    workbook.save("personnes.xlsx")
    exit()


def sup_tele():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active

    def delt():
        nation = str(input("donner numéro tele de personne désiré a supprimé :"))
        for i in range(1, ltab() + 1):
            if sheet["D" + str(i)].value == nation:
                sheet.delete_rows(i)

    delt()
    workbook.save("personnes.xlsx")
    exit()


def modifier_tele():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    tele = input("donner ancien num de personne a modifié :")
    for i in range(1, ltab() + 1):
        if sheet["D" + str(i)].value == tele:
            sheet["D" + str(i)].value = input("donner noveau num tele :")
    workbook.save("personnes.xlsx")
    exit()


def modifier_adr():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    tele = input("donner ancien adresse de personne a modifié :")
    for i in range(1, ltab() + 1):
        if sheet["F" + str(i)].value == tele:
            sheet["F" + str(i)].value = input("donner noveau adresse :")
    workbook.save("personnes.xlsx")
    exit()


def aff():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    for i in range(1, ltab() + 1):
        for j in range(ord("A"), ord("L")):
            if j == ord("A") and i > 1:
                print(sheet[chr(j) + str(i)].value, end=" ")
            else:
                print(sheet[chr(j) + str(i)].value, end="\t")
        print()


def aff_tele():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    tele = input("donner num tele de personne a rechercher :")
    for i in range(1, ltab() + 1):
        if i == 1:
            print(sheet[chr(j) + str(i)].value, end="\t")
        elif tele == sheet["D" + str(i)].value:
            print(sheet[chr(j) + str(i)].value, end="\t")
    exit()


def aff_cin():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    cin = input("donner le cin de personne désiré chercher :")
    for i in range(1, ltab() + 1):
        for j in range(ord("A"), ord("L")):
            if i == 1:
                print(sheet[chr(j) + str(i)].value, end="\t")
            elif sheet["A" + str(i)] == cin:
                print(sheet[chr(j) + str(i)].value, end="  ")
        print()
    exit()


def aff_nation():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    nation = input("donner la nationalité désiré afficher :")
    for i in range(1, ltab() + 1):
        for j in range(ord("A"), ord("K")):
            if i == 1:
                print(sheet[chr(j) + str(i)].value, end="\t")
            elif sheet["E" + str(i)].value == nation:
                if chr(j) == "A":
                    print(sheet[chr(j) + str(i)].value, end=" ")
                else:
                    print(sheet[chr(j) + str(i)].value, end="\t")
        print()
    exit()


def aff_décés():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    for i in range(1, ltab() + 1):
        for j in range(ord("A"), ord("K")):
            if i == 1:
                print(sheet[chr(j) + str(i)].value, end="\t")
            elif sheet["J" + str(i)] == 0:
                print(sheet[chr(j) + str(i)].value, end="  ")
        print()
    exit()


def aff_nondécés():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    for i in range(1, ltab() + 1):
        for j in range(ord("A"), ord("K")):
            if i == 1:
                print(sheet[chr(j) + str(i)].value, end="\t")
            elif sheet["J" + str(i)] == 1:
                print(sheet[chr(j) + str(i)].value, end="  ")
        print()
    exit()


def ajout_maladie():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active

    def ajouter():
        maladie = {
            "code": "",
            "cin": "",
            "Nom maladie": "",
            "nombreannee": 1,
        }
        for cle in maladie:
            print("donner ", cle, ":")
            maladie[cle] = input()
        sheet.append(
            [
                maladie["code"],
                maladie["cin"],
                maladie["Nom maladie"],
                maladie["nombreannee"],
            ]
        )

    ajouter()
    workbook.save("Maladies.xlsx")
    exit()


def sup_maladie():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active

    def delt():
        cin = int(input("donner code de maladie désiré a supprimé :"))
        cin = str(cin)
        for i in range(1, ltabm() + 1):
            if sheet["A" + str(i)].value == cin:
                sheet.delete_rows(i)

    delt()
    workbook.save("Maladies.xlsx")
    exit()


def nbannee_maladie():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active

    def mod():
        cin = int(input("donner code de maladie désiré a modifié :"))
        cin = str(cin)
        for i in range(2, ltabm() + 1):
            if sheet["A" + str(i)].value == cin:
                x = input("donner nb d'annee noveau :")
                sheet["D" + str(i)].value = x

    mod()
    workbook.save("Maladies.xlsx")
    exit()


def modif_décés():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active

    def mod():
        cin = int(input("donner code de maladie désiré a modifié :"))
        cin = str(cin)
        for i in range(1, ltabm() + 1):
            if sheet["A" + str(i)].value == cin:
                x = int(input("donner etat noveau:"))
                sheet["A" + str(i)].value = x

    mod()
    workbook.save("Maladies.xlsx")
    exit()


def aff_maladie():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active
    for i in range(1, ltabm() + 1):
        for j in range(ord("A"), ord("E")):
            if i == 1:
                print(sheet[chr(j) + str(i)].value, end="\t")
            else:
                if chr(j) == "B":
                    print(sheet[chr(j) + str(i)].value, end=" ")
                else:
                    print(sheet[chr(j) + str(i)].value, end="\t")
        print()
    exit()


def rech_maladie():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active
    code = input("donner code de maladie :")
    for i in range(1, ltabm() + 1):
        for j in range(ord("A"), ord("E")):
            if i == 1:
                print(sheet[chr(j) + str(i)].value, end="\t")
            elif sheet["A" + str(i)].value == code:
                print(sheet[chr(j) + str(i)].value, end="  ")
        print()
    exit()


def maladie_personne():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active
    cin = input("donner cin de personne :")
    for i in range(1, ltabm() + 1):
        for j in range(ord("A"), ord("E")):
            if i == 1:
                print(sheet[chr(j) + str(i)].value, end="\t")
            elif sheet["B" + str(i)].value == cin:
                print(sheet[chr(j) + str(i)].value, end="  ")
        print()
    exit()


def pourcentage():
    workbook = load_workbook("Maladies.xlsx")
    sheet = workbook.active
    l = []
    k = 1
    for i in range(1, ltabm() + 1):
        if sheet["C" + str(i)].value != l:
            r = str(sheet["C" + str(i)].value) + str(k)
            l.append(r)
        else:
            k = k + 1
            j = l.index(sheet["C" + str(i)].value)
            l[j] = sheet["C" + str(i)] + str(k)
    p = 0
    for i in l:
        p = int(i[len(l) - 2 :])
        print(str(i[len(l) - 2 :]) + str(len(l) / p * 100) + "%")
    exit()


def maladiechaque():
    workbook1 = load_workbook("personnes.xlsx")
    workbook2 = load_workbook("Maladies.xlsx")
    sheet1 = workbook1.active
    sheet2 = workbook2.active
    for i in range(0, ltabm() + 1):
        for j in range(ord("A"), ord("E")):
            if i == 1:
                print(sheet2[chr(j) + str(i)].value, end="\t")
            elif sheet2["A" + str(i)] == cin:
                print(sheet2[chr(j) + str(i)].value, end="  ")
        print()
    for i in range(0, ltab() + 1):
        for j in range(ord("A"), ord("K")):
            print(sheet1[chr(j) + str(i)].value, end="\t")
        print()
    exit()


def quar():
    workbook = load_workbook("personnes.xlsx")
    sheet = workbook.active
    today = date.today()
    today = str(today)
    year = int(today[:4])
    month = int(today[6:7])
    day = int(today[9:11])
    for i in range(1, 2):
        for j in range(ord("A"), ord("L")):
            print(sheet[chr(j) + str(i)].value, end="\t")
        print()
    for i in range(2, ltab() + 1):
        if (year * 1000000 + month * 100 + day) - (
            int(sheet["J" + str(i)].value) * 1000000
            + int(sheet["I" + str(i)].value) * 100
            + int(sheet["H" + str(i)].value)
        ) <= 14:
            for j in range(ord("A"), ord("L")):
                print(sheet[chr(j) + str(i)].value, end="\t")
            print()
    exit()


def do_something():
    print("button clicked")


# Create the main window
root = tk.Tk()
bgimg = tk.PhotoImage(file="img.ppm")
limg = Label(root, i=bgimg)
limg.pack()
root.geometry("600x300")
# Create a menu bar
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)
# Create a File menu with some items
file_menu = tk.Menu(menu_bar, tearoff=0)
rech_menu = tk.Menu(file_menu, tearoff=0)
file_menu.add_cascade(label="Menu Personnes", menu=rech_menu)
as_menu = tk.Menu(rech_menu, tearoff=0)
rech_menu.add_cascade(label="Mise a jour Personnes", menu=as_menu)
as_menu.add_command(label="Ajouter personne", command=ajout_pers)
ass_menu = tk.Menu(rech_menu, tearoff=0)
as_menu.add_cascade(label="Supprimer personne", menu=ass_menu)
ass_menu.add_command(label="Suppression personne donné", command=sup_pers)
ass_menu.add_command(
    label="Suppression des personnes d'une nationalité donnée", command=sup_nation
)
ass_menu.add_command(
    label="Suppression des personnes d'un indicatif donnée (téléphone)",
    command=sup_tele,
)
asss_menu = tk.Menu(rech_menu, tearoff=0)
as_menu.add_cascade(label="Modifier personne", menu=asss_menu)
asss_menu.add_command(label="telephone", command=modifier_tele)
asss_menu.add_command(label="Adresse", command=modifier_adr)
recherche_menu = tk.Menu(rech_menu, tearoff=0)
rech_menu.add_cascade(label="rechercher,afficher", menu=recherche_menu)
recherche_menu.add_command(label="contenu du dictionnaire personne", command=aff)
recherche_menu.add_command(label="recherche par numéro téléphone", command=aff_tele)
recherche_menu.add_command(label="recherche par indicatif", command=aff_cin)
recherche_menu.add_command(label="recherche par nationalité", command=aff_nation)
recherche_menu.add_command(label="recherche des personnes décédés", command=aff_décés)
recherche_menu.add_command(
    label="recherche des personnes non décédés", command=aff_nondécés
)
menu_bar.add_cascade(label="Personnes", menu=file_menu)

# Create an Edit menu with some items
fil_menu = tk.Menu(menu_bar, tearoff=0)
mise_menu = tk.Menu(menu_bar, tearoff=0)
aff_menu = tk.Menu(menu_bar, tearoff=0)
rech1_menu = tk.Menu(menu_bar, tearoff=0)
rech2_menu = tk.Menu(menu_bar, tearoff=0)
fil_menu.add_cascade(label="Mise a jour", menu=mise_menu)
mise_menu.add_command(label="Ajouter une nouvelle maladie", command=ajout_maladie)
mise_menu.add_command(label="supprimer une maladie", command=sup_maladie)
mise_menu.add_cascade(label="Modfifier les données d'une maladie", menu=aff_menu)
aff_menu.add_command(label="Nombre d'années", command=nbannee_maladie)
aff_menu.add_command(label="Modifier décés (de 0 a 1)", command=modif_décés)
fil_menu.add_cascade(label="Recherche,affichage", menu=rech1_menu)
rech1_menu.add_command(label="Contenu du dictionnaire maladies", command=aff_maladie)
rech1_menu.add_command(label="Recherche par une maladie", command=rech_maladie)
rech1_menu.add_command(
    label="Recherche maladies d'une personne", command=maladie_personne
)
rech1_menu.add_command(
    label="Recherche le pourcentage de chaque maladie", command=pourcentage
)
rech1_menu.add_command(
    label="Recherche maladies de chaque personne", command=do_something
)
menu_bar.add_cascade(label="Maladies", menu=fil_menu)
menu_bar.add_cascade(label="Calcul et affichage", menu=rech2_menu)
rech2_menu.add_command(label="Afficher par nationalité", command=do_something)
rech2_menu.add_command(label="personnes en quarantine", command=quar)
rech2_menu.add_command(label="personnes décés", command=do_something)
rech2_menu.add_command(label="personnes à risque", command=do_something)

# Start the main loop
root.mainloop()
